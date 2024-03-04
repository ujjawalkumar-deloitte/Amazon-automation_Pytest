import openpyxl


def read_excel(file_path, return_test):

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    base_url = sheet.cell(row=2, column=1).value
    email = str(sheet.cell(row=2, column=2).value)
    password = sheet.cell(row=2, column=3).value
    full_name = sheet.cell(row=2, column=4).value
    mobile_number = str(sheet.cell(row=2, column=5).value)
    pincode = sheet.cell(row=2, column=6).value
    house_no = sheet.cell(row=2, column=7).value
    area = sheet.cell(row=2, column=8).value
    landmark = sheet.cell(row=2, column=9).value
    search = sheet.cell(row=2, column=10).value

    workbook.close()

    if return_test == 1:
        return base_url, email, password
    elif return_test == 4:
        return full_name, mobile_number, pincode, house_no, area, landmark
    elif return_test == 5:
        string = search
        words = string.split()
        return words[0]
        # return search[0]
    elif return_test == 3:
        string = search
        words = string.split()
        # print(words)
        return words

    # if __name__ == "__main__":
    #     path = r"test_Data\assignment_data.xlsx"
    #
    #     words_list = read_excel(path, 3)
    #     print(words_list)
    #     for item in words_list:
    #         print(f"This is item: {item}")
    #
